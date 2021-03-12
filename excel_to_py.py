from openpyxl import Workbook
from openpyxl import load_workbook
import re
print('excel to py file loaded...')

file = input('whats the excelfile name?(.xlsm) type "default" for est model')
if file == 'default':
    file = 'Autowrite excel model1.0.6.xlsm'

workbook = load_workbook(filename=file,data_only=True)

estimate_total_sheet = workbook['Estimate Total']
project_materials_sheet =workbook['Project Materials']
area_a_sheet = workbook['Area A']
area_b_sheet =workbook['Area B']
area_c_sheet = workbook['Area C']
area_d_sheet = workbook['Area D']
grab_rail_sheet = workbook['Grab Rail']
project_labor_sheet = workbook['Project Labor']
take_off_sheet = workbook['Take Off']
item_list_sheet = workbook['Item List']
grab_rail_item_sheet = workbook['Grab Rail Item List']
note_sheet = workbook['Write up']

estimate = []


def return_variables():
    for row in note_sheet.iter_rows(min_row=1,min_col=2,max_col=2,values_only=True):
        res = ''.join(map(str,row))
        estimate.append(res)
    return estimate[0],estimate[1],estimate[2],estimate[3],estimate[4],estimate[5]

def return_lf():
    section_lf= []
    for row in note_sheet.iter_rows(min_row=2,min_col=4,max_col=4,values_only=True):
        res = ''.join(map(str,row))
        if res != "None":
            if res != "0":
                res = round(float(res),0)
                section_lf.append(int(res))
    return section_lf

def return_lfprice():
    section_lfprice= []
    for row in note_sheet.iter_rows(min_row=2,min_col=5,max_col=5,values_only=True):
        res = ''.join(map(str,row))
        if res != 'None':
            if res != '0':
                res = round(float(res),0)
                section_lfprice.append(int(res))
    return section_lfprice

def return_section_details(num=0):
    sections = return_lf()
    section = []
    for row in note_sheet.iter_rows(min_row=2,min_col=(7+num),max_col=(7+num),values_only=True):
            res = ''.join(map(str,row))
            if res != 'None':
                section.append(int(res))
    return section

def return_area_name(num=0):
    area_names = ''
    for row in note_sheet.iter_rows(min_row=(2+num),max_row=(2+num),min_col=3,max_col=3,values_only=True):
        print(row)
        res = ''.join(map(str,row))
        if res != 'None':
            area_name = str(res)
            return area_name
        else:
            return 'None'

def return_rep():
    for row in note_sheet.iter_rows(min_row=13,max_row=13,min_col=2,max_col=2,values_only=True):
        rep = ''.join(map(str,row))
        if rep.lower() == 'jag':
            return 'jag'
        elif rep.lower() == 'dave':
            return 'dave'
        else:
            return 'jag'
    

