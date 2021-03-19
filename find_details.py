from openpyxl import Workbook
from openpyxl import load_workbook
import excel_to_py as etp

file = etp.file

wb= load_workbook(file,data_only=True)

top_rail=''
bottom_rail=''
infill=''
mounting=''

area_a = wb['Area A']

for item in area_a.iter_rows(min_row=5,min_col=2,max_col=2,max_row=5,values_only=True):
        print(item)
        if 'TR200' in str(item):
                top_rail = 0
        elif 'TR375' in str(item):
                top_rail = 1

print(top_rail)

for item in area_a.iter_rows(min_row=26,max_row=36,min_col=2,max_col=2,values_only=True):
        print(item)
        if 'PT' in str(item):
                infill=0
        elif 'CTLG' in str(item):
                infill=1
        elif 'Cable' in str(item):
                infill=2
print(infill)
