from openpyxl import Workbook
from openpyxl import load_workbook
import excel_to_py as etp

file = etp.file

wb= load_workbook(file,data_only=True)

top_rail=''
bottom_rail=''
infill=''
mounting=''
post_height=''


area_a = wb['Area A']

for item in area_a.iter_rows(min_row=5,min_col=2,max_col=2,max_row=5,values_only=True):
        print(item)
        if 'TR200' in str(item):
                top_rail = 0
        elif 'TR375' in str(item):
                top_rail = 1
        elif 'TR400' in str(item):
                top_rail=2
        else:
                top_rail=8
print(top_rail)

for item in area_a.iter_rows(min_row=26,max_row=36,min_col=2,max_col=2,values_only=True):
        print(item)
        if 'PT' in str(item):
                infill=0
        elif 'CTLG' in str(item):
                infill=1
        elif 'Cable' in str(item):
                infill=2
print(infill)

for item in area_a.iter_rows(min_row=12,max_row=18,min_col=2,max_col=2,values_only=True):
        print(item)
        if 'FP56' in str(item):
                post_height= 1
        elif 'FP45' in str(item):
                post_height=0
        elif '36' in str(item):
                post_height=0
                mounting=3
        elif '42' in str(item):
                post_height=1
                mounting=3
        else:
                post_height=3
print(post_height)

for item in area_a.iter_rows(min_row=22,max_row=25,max_col=2,min_col=2,values_only=True):
        print(item)
        if 'FMPBS 1' in str(item):
                mounting=0
        elif 'BP-' in str(item):
                mounting=3
        elif 'BP Stairs' in str(item):
                mounting=6
        elif 'L Brackets' in str(item):
                mounting=5
print(mounting)
                
